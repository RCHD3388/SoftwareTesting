import openpyxl
import os

file_path = 'configuration/RequestInstructor.xlsx'
book = openpyxl.load_workbook(file_path)
sheet = book.active

print("Current headers:")
headers = []
for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    headers.append(header)
print(headers)

print("\nExpected headers:")
expected_headers = ['test_case_name', 'firstName', 'lastName', 'phoneNumber', 'address', 'cv', 'bio', 'title']
print(expected_headers)

print("\nFirst row data:")
data = []
for col in range(1, sheet.max_column + 1):
    value = sheet.cell(row=2, column=col).value
    data.append(value)
print(data)
