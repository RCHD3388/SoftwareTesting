import openpyxl
import os

dir_path = os.path.dirname(os.path.realpath(__file__))

class PageData:
    @staticmethod
    def getTestData(placeholder_data, test_case_name):
        file_path = os.path.join(dir_path, f'../configuration/{placeholder_data}.xlsx')
        
        Dict = {}
        book = openpyxl.load_workbook(file_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(list(Dict.values()))

        
        return list(Dict.values())

